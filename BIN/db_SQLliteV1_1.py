import pandas as pd
from tkinter import filedialog
from sqlalchemy import create_engine
import os
import csv
import datetime
import openpyxl
import re
run_path=os.getcwd()
out_path=os.path.join(str(run_path),"OUTPUT")
in_path=os.path.join(run_path, "INPUT")
config_path=os.path.join(run_path, "CONFIG")
bin_path=os.path.join(run_path, "BIN")
#Create folders if not exist
if not os.path.exists(out_path):    os.makedirs(out_path)
if not os.path.exists(in_path):    os.makedirs(in_path)
if not os.path.exists(config_path):    os.makedirs(config_path)
if not os.path.exists(bin_path):    os.makedirs(bin_path)

# Connect to the database
engine = create_engine('sqlite:///BIN/mydatabase.db')

def load_data_from_files(table = "PACE_TABLE",file_path = "",sheet_name=0,action="replace"):
    if file_path == "":
        #open file explorer 
        file_path = filedialog.askopenfilename(title="Select input file", filetypes=(("excel files", "*.xlsx"), ("csv files", "*.csv"),("all files", "*.*")),initialdir=run_path,multiple=True)
    if not isinstance(file_path, list):
        file_path = [file_path]
    with engine.connect() as con:
        con.exec_driver_sql("drop table if exists " + table)
    for file in file_path:
        # fill db from exel file
        #print("reading file "+file)
        ext = file[-4:].upper()
        if ext == ".CSV":
            df = pd.read_csv(file)
        elif ext == ".XLS":
            df = pd.read_excel(file,engine='openpyxl')
        elif ext == "XLSX":
            df = pd.read_excel(file,sheet_name=sheet_name,engine='openpyxl')
        else:
            print("file format not supported")
            return False    
        print("creating table: "+ table +" in DB from: " + os.path.basename(file)  )
        df.to_sql(table, engine,if_exists= action ,index=False)
    return True

# Convert the DataFrame to a SQL table


def create_file(full_file_text,file_type='xlsx'):
    

    for file_text in full_file_text.split('---separator---'):
        if file_text.strip():
            #print(file_text)
            file_name = re.search('----File:(.+?)----', file_text).group(1)
            file_name = file_name.replace("MM-DD-YYYY", datetime.date.today().strftime("%m-%d-%Y"))

            if file_name == "":
                file_name = "Export_" + datetime.date.today().strftime("%Y-%m-%d_%H%M") + "."+file_type
            #print(file_name)

            try:
                writer = pd.ExcelWriter(os.path.join(out_path,file_name), engine='openpyxl',date_format='MM-DD-YYYY',datetime_format='MM-DD-YYYY')
                
                for statement in file_text.split(';'):
                    if statement.strip():
                        sheet_name = re.search('----Sheet:(.+?)----', statement)
                        if sheet_name is not None:
                            sheet_name = sheet_name.group(1)
                        if sheet_name == "" or sheet_name is None:
                            sheet_name = "Sheet1"

                        headers_table = re.search('----Header:(.+?)----', statement)
                        if headers_table is not None:
                            headers_table = headers_table.group(1)
                        
                        if headers_table != "" and headers_table is not None:
                            headers_statement = "select * from " + headers_table
                            try:
                                with engine.connect() as con:
                                    headers = con.exec_driver_sql(headers_statement)
                            except Exception as e:
                                print ("------------------Exeption on headers query------------------")
                                print(e)
                                print ("-----------------------------------------------------")
                                return False
                            
                            
                            headers_df = pd.DataFrame(headers.all())
                            #check df elements length
                            
                            if len(headers_df.columns) > 0:
                                headers_df.columns = headers.keys()
                                list_columns = list(headers.keys())
                                single_row = False
                            else:
                                del headers_df
                                headers_df = pd.DataFrame(headers.keys())
                                list_columns = list(headers.keys())
                                #Print(list_columns)
                                single_row = True
                        print("creating sheet: "+ sheet_name +" in file: " + file_name  )
                        try:
                            with engine.connect() as con:
                                tabla_custom = con.exec_driver_sql(statement)
                        except Exception as e:
                            print ("------------------Exeption on query------------------")
                            print(e)
                            print ("-----------------------------------------------------")
                            return False
                                            
                        try:
                            df = pd.DataFrame(tabla_custom.all())
                        except Exception as e:
                            print("-----------------error reading data from db-----------------")
                            print(e)
                            print("------------------------------------------------------------")
                            return False
                        df.columns = tabla_custom.keys()
                                    
                        r = re.compile('\d{2}-\d{2}-\d{4} *') 
                        r2 = re.compile('\d{4}-\d{2}-\d{2} *')
                        r3 = re.compile('\d{2}/\d{2}/\d{4}')
                        r4 = re.compile('\d{2}/\d{1}/\d{4}')
                        r5 = re.compile('\d{1}/\d{1}/\d{4}')
                        
                        for (column, data) in df.items():  
                            #df[column] = df[column].str.replace('', '') #remove strange character
                            for row in data.values:
                                try:
                                    if r.match(row) is not None or r2.match(row) is not None or r3.match(row) is not None or r4.match(row) is not None or r5.match(row) is not None:
                                        try:
                                            df[column] = pd.to_datetime(df[column])
                                            break
                                        except ValueError:
                                            break
                                    else:
                                        try:
                                            df[column] = pd.to_numeric(df[column])
                                            break
                                        except ValueError:
                                            break
                                except:
                                    pass
                        try:
                            df.to_excel(writer, index=False,sheet_name=sheet_name,float_format="%.2f")
                        except Exception as e:
                            print("-----------------error writing data to file-----------------")
                            print(e)
                            print("------------------------------------------------------------")
                            continue
                        worksheet = writer.sheets[sheet_name] #Get the xlsxwriter worksheet object
                        #Write the column headers with the defined format
                        #__________________________________________________________________________________________________________________________________
                        def get_column_letter(col_num, start=1):
                            """Return the column letter for the given column number.For example, 1 -> 'A', 2 -> 'B', 27 -> 'AA'."""
                            quot, rem = divmod(col_num - start, 26)
                            if quot > 0:     return get_column_letter(quot, start) + chr(rem + ord('A'))
                            else:            return chr(rem + ord('A'))
                        def get_col_widths(dataframe):
                            #First we find the maximum length of the index column   
                            idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
                            #Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
                            return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]
                        #__________________________________________________________________________________________________________________________________
                        for col_num, value in enumerate(df.columns.values):
                            #worksheet.write(0, col_num, value, header_format)
                            col = get_column_letter(col_num+1) + '1'
                            worksheet[col].font = openpyxl.styles.Font(bold=True, size=12, color='00FFFFFF', name='Calibri')
                            worksheet[col].fill = openpyxl.styles.PatternFill("solid", fgColor="366092")
                            worksheet[col].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center') #, wrap_text=True)
                            #'valign': 'top','align': 'center', 'text_wrap': True,'fg_color': '#366092', 'font_color': 'white', 'border': 1
                        
                        #Adjust every cell width
                        for col_num, width in enumerate(get_col_widths(df)):
                            #worksheet.set_column(col_num-1, col_num, width) 
                            worksheet.column_dimensions[get_column_letter(col_num+1)].width = 20 #width
                        #set datetime columns format
                        
                        for col_num, value in enumerate(df.columns.values):
                            if df[value].dtype == 'datetime64[ns]':
                                col = get_column_letter(col_num+1)
                                #iter over rows
                                for row in range(2,df.shape[0]+2):
                                    worksheet[col+str(row)].number_format = 'MM/DD/YYYY'
                                pass
                        #set headers
                        #insert lines in the first row
                        if headers_table != "" and headers_table is not None:
                            #inserline in excel
                            worksheet.insert_rows(1)
                            #insert headers
                            
                            for col_num, value in enumerate(list_columns):
                                col = get_column_letter(col_num+1) + '1'
                                if not  bool(re.match(r"Unnamed*",value)):
                                    worksheet[col] = value
                                worksheet[col].font = openpyxl.styles.Font(bold=False, size=12, color='00FFFFFF', name='Calibri')
                                worksheet[col].fill = openpyxl.styles.PatternFill("solid", fgColor="50b6fa")
                                worksheet[col].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                            #iter header_df data values
                            if single_row == False:
                                for i in headers_df.index:
                                    #insert line
                                    worksheet.insert_rows(i+2)
                                    #insert data
                                    for col_num, value in enumerate(headers_df.columns.values):
                                        col = get_column_letter(col_num+1) + str(i+2)
                                        worksheet[col] = headers_df[value][i]
                                        worksheet[col].font = openpyxl.styles.Font(bold=False, size=12, color='00FFFFFF', name='Calibri')
                                        worksheet[col].fill = openpyxl.styles.PatternFill("solid", fgColor="50b6fa")
                                        worksheet[col].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                            
                        
                writer.close()
                #print("file created")              
            except IOError:
                print("file is open")
                input("close file and press enter")
            
            
            
if __name__=="__main__":
    
    load_data_from_files()
    #load query from sql file
    with(open(os.path.join(config_path,"jessica_tam.sql"),'r')) as f:
        query=f.read()
    create_file(query)
    print("Complete")